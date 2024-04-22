import openpyxl
from datetime import datetime

sheetPath = '/home/szabo/Documents/Ficha Frequencia 2024 - Henrique Szabo.xlsx'

workbook = openpyxl.load_workbook(sheetPath)

sheet = workbook['ABR']

local_date = datetime.now().strftime('%d/%m/%Y')

for row in sheet.iter_rows(3, sheet.max_row):
    data_celula = row[0].value
    if isinstance(data_celula, datetime):
        data_celula = data_celula.strftime('%d/%m/%Y')
    if data_celula == local_date:
        row[1].value = 'Reunião de alinhamento + Oganização de tarefas com o time'
        row[4].value = 6

workbook.save(sheetPath)

workbook.close()
